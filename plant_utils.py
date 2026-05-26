"""Utilities for Plant Management, WhatsApp Notifications, and Excel Export"""
import os
from datetime import datetime
from pathlib import Path
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def send_whatsapp_notification(phone_number: str, message: str) -> bool:
    """
    Send WhatsApp notification via Twilio or similar service.
    Configure WHATSAPP_API_KEY, WHATSAPP_ACCOUNT_SID, WHATSAPP_PHONE in env vars.
    """
    try:
        account_sid = os.getenv("WHATSAPP_ACCOUNT_SID")
        auth_token = os.getenv("WHATSAPP_AUTH_TOKEN")
        twilio_phone = os.getenv("WHATSAPP_PHONE", "whatsapp:+14155552671")
        
        if not account_sid or not auth_token:
            print(f"⚠️ WhatsApp no configurado. Mensaje para {phone_number}: {message}")
            return False
        
        # Usar Twilio API
        url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
        
        data = {
            "From": twilio_phone,
            "To": f"whatsapp:+{phone_number}" if not phone_number.startswith("whatsapp") else phone_number,
            "Body": message
        }
        
        response = requests.post(
            url,
            data=data,
            auth=(account_sid, auth_token),
            timeout=10
        )
        
        if response.status_code in [201, 200]:
            print(f"✅ WhatsApp enviado a {phone_number}")
            return True
        else:
            print(f"❌ Error al enviar WhatsApp: {response.text}")
            return False
            
    except Exception as e:
        print(f"❌ Error enviando WhatsApp: {str(e)}")
        return False


def get_priority_order(priority: str) -> int:
    """Get priority order for sorting (1=emergencia, 2=alta, 3=media, 4=baja)"""
    priority_map = {
        "emergencia": 1,
        "alta": 2,
        "media": 3,
        "baja": 4
    }
    return priority_map.get(priority.lower(), 5)


def export_tickets_to_excel_multiple_sheets(tickets: list, plant_name: str, output_path: str = "reports/tickets_export.xlsx") -> str:
    """
    Export tickets to Excel with 4 sheets (Completadas, Incompletas, En Proceso, Canceladas)
    Each sheet has its own color and tickets are sorted by priority.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define sheet configurations
    sheets_config = {
        "Completadas": {
            "color": "c6efce",  # Light green
            "header_color": "10b981",  # Dark green
            "filter_statuses": ["completo", "cerrado"]
        },
        "Incompletas": {
            "color": "ffc7ce",  # Light red
            "header_color": "ef4444",  # Dark red
            "filter_statuses": ["incompleto", "abierto", "pendiente"]
        },
        "Neutras": {
            "color": "dbeafe",  # Light blue
            "header_color": "3b82f6",  # Blue
            "filter_statuses": ["proceso", "en_proceso", "abierto", "pendiente"]
        },
        "Canceladas": {
            "color": "e5e7eb",  # Light gray
            "header_color": "6b7280",  # Dark gray
            "filter_statuses": ["cancelado", "rechazado", "cancelled"]
        }
    }
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["ID", "Asunto", "Estado", "Prioridad", "Responsable", "Planta", "Fecha Creación", "Descripción"]
    
    # Priority mapping for colors
    priority_colors = {
        "emergencia": "fccccb",  # Light red for emergency
        "alta": "fce4d6",        # Light orange for high
        "media": "e2efda",       # Light green for medium
        "baja": "d9e1f2"         # Light blue for low
    }
    
    # Create sheets
    for sheet_name, config in sheets_config.items():
        ws = wb.create_sheet(sheet_name)
        
        # Filter tickets by status
        filtered_tickets = [t for t in tickets if t.status.lower() in config["filter_statuses"]]
        
        # Sort by priority (emergencia > alta > media > baja)
        filtered_tickets.sort(key=lambda t: get_priority_order(t.priority))
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color=config["header_color"], end_color=config["header_color"], fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Create headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data rows
        for row_num, ticket in enumerate(filtered_tickets, 2):
            # Get priority color
            priority_color = priority_colors.get(ticket.priority.lower(), "ffffff")
            fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type="solid")
            
            data = [
                ticket.id,
                ticket.subject,
                ticket.status,
                ticket.priority,
                ticket.assigned_employee.name if ticket.assigned_employee else "Sin asignar",
                plant_name,
                ticket.created_at.strftime("%d/%m/%Y %H:%M") if ticket.created_at else "",
                ticket.description[:50] + "..." if len(ticket.description) > 50 else ticket.description
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Adjust column widths
        column_widths = [8, 30, 15, 12, 20, 15, 18, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    # Create output directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    return output_path


def export_tickets_to_excel(tickets: list, plant_name: str, output_path: str = "reports/tickets_export.xlsx") -> str:
    """
    Export tickets to Excel with color coding by status.
    Colors: Verde=Completo, Naranja=En proceso, Rojo=Incompleto
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Define colors
    colors = {
        "completo": "10b981",      # Green
        "proceso": "f59e0b",        # Orange
        "incompleto": "ef4444",     # Red
    }
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["ID", "Asunto", "Estado", "Prioridad", "Responsable", "Planta", "Fecha Creación", "Descripción"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    status_colors = {
        "completo": "10b981",
        "proceso": "f59e0b",
        "incompleto": "ef4444",
        "abierto": "3b82f6",
        "cerrado": "10b981",
    }
    
    for row_num, ticket in enumerate(tickets, 2):
        status = getattr(ticket, 'status', 'desconocido').lower()
        color = status_colors.get(status, "d1d5db")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        data = [
            ticket.id,
            ticket.subject,
            ticket.status,
            ticket.priority,
            ticket.assigned_employee.name if ticket.assigned_employee else "Sin asignar",
            plant_name,
            ticket.created_at.strftime("%d/%m/%Y %H:%M") if ticket.created_at else "",
            ticket.description[:50] + "..." if len(ticket.description) > 50 else ticket.description
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = [8, 30, 15, 12, 20, 15, 18, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create output directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    return output_path


def calculate_ticket_stats(tickets: list) -> dict:
    """Calculate ticket statistics for dashboard"""
    total = len(tickets)
    completo = sum(1 for t in tickets if t.status.lower() in ["completo", "cerrado"])
    incompleto = sum(1 for t in tickets if t.status.lower() in ["incompleto", "abierto"])
    neutro = sum(1 for t in tickets if t.status.lower() in ["proceso", "en_proceso", "pendiente"])
    cancelado = sum(1 for t in tickets if t.status.lower() in ["cancelado", "rechazado", "cancelled"])
    
    # Any ticket that does not fit explicit categories is treated as neutral
    otros = total - (completo + incompleto + neutro + cancelado)
    neutro += otros
    
    return {
        "total": total,
        "completo": completo,
        "incompleto": incompleto,
        "neutro": neutro,
        "cancelado": cancelado,
        "completo_percent": round((completo / total * 100) if total > 0 else 0),
        "incompleto_percent": round((incompleto / total * 100) if total > 0 else 0),
        "neutro_percent": round((neutro / total * 100) if total > 0 else 0),
        "cancelado_percent": round((cancelado / total * 100) if total > 0 else 0),
    }


def filter_tickets_by_status(tickets: list, status: str) -> list:
    """Filter tickets by status"""
    status_lower = status.lower()
    if status_lower == "completo":
        return [t for t in tickets if t.status.lower() in ["completo", "cerrado"]]
    elif status_lower == "incompleto":
        return [t for t in tickets if t.status.lower() in ["incompleto", "abierto"]]
    elif status_lower == "proceso":
        return [t for t in tickets if t.status.lower() in ["proceso", "en_proceso"]]
    return []

